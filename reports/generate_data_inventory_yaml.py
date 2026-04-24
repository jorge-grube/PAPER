from pathlib import Path
import re
import pandas as pd
import yaml
from openpyxl import load_workbook

ROOT = Path(r"c:/Users/Jorge/OneDrive - UFV/BAINF/PAPER")
DATA_ROOT = ROOT / "data"
OUT_PATH = ROOT / "reports" / "data_inventory.yml"
FILES = sorted(DATA_ROOT.rglob("*.xlsx"))

INTERVAL_MAP = {"P1D": "daily", "P1W": "weekly", "P1M": "monthly", "P3M": "quarterly", "P6M": "semiannual", "P1Y": "annual"}


def norm(value):
    if value is None:
        return ""
    text = str(value).strip()
    return "" if text.lower() in {"none", "nan", "nat"} else text


def to_date(value):
    if value is None:
        return None
    if hasattr(value, "year") and hasattr(value, "month") and hasattr(value, "day"):
        return pd.Timestamp(value).to_pydatetime()
    text = norm(value)
    if not text:
        return None
    for dayfirst in (True, False):
        ts = pd.to_datetime(text, errors="coerce", dayfirst=dayfirst)
        if pd.notna(ts):
            return pd.Timestamp(ts).to_pydatetime()
    try:
        num = float(text.replace(",", "."))
        if 20000 <= num <= 60000:
            return pd.to_datetime(num, unit="D", origin="1899-12-30").to_pydatetime()
    except Exception:
        pass
    return None


def to_num(value):
    if value is None:
        return None
    if isinstance(value, (int, float)):
        return float(value)
    text = norm(value).replace("\u00a0", "").replace("%", "")
    if not text:
        return None
    if text.count(",") == 1 and text.count(".") == 0:
        text = text.replace(",", ".")
    elif text.count(",") and text.count("."):
        if text.rfind(",") > text.rfind("."):
            text = text.replace(".", "").replace(",", ".")
        else:
            text = text.replace(",", "")
    try:
        return float(text)
    except Exception:
        return None


def infer_unit(rel_path):
    low = rel_path.lower()
    name = Path(rel_path).name.lower()
    if "/investable_assets/equity/" in low or "/investable_assets/commodities/" in low or "/investable_assets/real_assets/" in low:
        return "price_index_level"
    if "/investable_assets/cash/" in low:
        return "rate_level"
    if "/regime_variables/volatility/" in low:
        return "volatility_index_level"
    if "/regime_variables/sovereign_yields/" in low or "yield" in name:
        return "yield_level"
    if "hicp" in name or "ppi" in name:
        return "inflation_level"
    if "pmi" in name:
        return "diffusion_index"
    if "confidence" in name or "sentiment" in name or "zew" in name:
        return "sentiment_index"
    if "gdp" in name or "unemployment" in name or "industrial_production" in name:
        return "macro_level"
    if "balance_sheet" in name:
        return "balance_sheet_level"
    if "dxy" in name:
        return "fx_index_level"
    return "level"


def infer_freq_from_dates(dates):
    series = pd.Series([pd.Timestamp(d) for d in dates if d is not None]).dropna().drop_duplicates().sort_values()
    if len(series) < 3:
        return "unknown"
    gaps = series.diff().dropna().dt.days
    if gaps.empty:
        return "unknown"
    med = float(gaps.median())
    if med <= 2:
        return "daily"
    if med <= 8:
        return "weekly"
    if med <= 16:
        return "biweekly"
    if med <= 40:
        return "monthly"
    if med <= 100:
        return "quarterly"
    if med <= 370:
        return "annual"
    return "irregular"


def infer_freq_from_formula(formulas):
    joined = " ".join(str(f) for f in formulas if f)
    for token, label in INTERVAL_MAP.items():
        if token in joined:
            return label
    return "unknown"


def clean_variables(values):
    out = []
    for value in values:
        text = norm(value)
        if not text or re.match(r"^Unnamed:", text):
            continue
        out.append(text)
    return out


def find_formula_anchor(formulas):
    joined = " ".join(formulas)
    m = re.search(r"\$A\$(\d+)", joined)
    return int(m.group(1)) if m else None


def detect_sheet(fp):
    try:
        wb_data = load_workbook(fp, data_only=True, read_only=True)
        wb_formula = load_workbook(fp, data_only=False, read_only=True)
    except Exception as exc:
        return {
            "sheet": None,
            "header_row": None,
            "data_start_row": None,
            "freq": None,
            "variables": [],
            "status": "UNREADABLE",
            "notes": f"cannot open workbook: {exc}",
            "engine": "openpyxl",
        }

    best = None
    for ws_d, ws_f in zip(wb_data.worksheets, wb_formula.worksheets):
        data_start_row = None
        mode = "none"

        # 1) cached real observations
        for r in range(1, min(ws_d.max_row, 200) + 1):
            vals = [ws_d.cell(r, c).value for c in range(1, min(ws_d.max_column, 12) + 1)]
            if any(to_date(v) is not None for v in vals[:3]) and sum(to_num(v) is not None for v in vals[1:]) >= 1:
                data_start_row = r
                mode = "cached"
                break

        # 2) formula anchor for sheets that require refresh
        formulas = []
        if data_start_row is None:
            for r in range(1, min(ws_f.max_row, 60) + 1):
                for c in range(1, min(ws_f.max_column, 20) + 1):
                    v = ws_f.cell(r, c).value
                    if isinstance(v, str) and v.startswith("="):
                        formulas.append(v)
            if formulas:
                anchor = find_formula_anchor(formulas)
                if anchor is not None:
                    data_start_row = anchor
                    mode = "formula_anchor"
                elif any("HistoricalPricing" in f or "RDP.HistoricalPricing" in f or "_xll.RDP.HistoricalPricing" in f for f in formulas):
                    data_start_row = min(r for r in range(1, min(ws_f.max_row, 60) + 1) if any(isinstance(ws_f.cell(r, c).value, str) and ws_f.cell(r, c).value.startswith("=") for c in range(1, min(ws_f.max_column, 20) + 1))) + 1
                    mode = "formula_generic"

        if data_start_row is None:
            continue

        header_row = data_start_row - 1
        while header_row > 1 and not any(norm(ws_f.cell(header_row, c).value) for c in range(1, min(ws_f.max_column, 30) + 1)):
            header_row -= 1

        variables = clean_variables(ws_f.cell(header_row, c).value for c in range(1, min(ws_f.max_column, 30) + 1))
        if not variables:
            continue

        if mode == "cached":
            dates = []
            for r in range(data_start_row, min(ws_d.max_row, data_start_row + 6000) + 1):
                d = to_date(ws_d.cell(r, 1).value)
                if d is not None:
                    dates.append(d)
            freq = infer_freq_from_dates(dates)
            status = "OK"
            notes = f"mode=cached rows={ws_d.max_row} cols={ws_d.max_column}"
        else:
            freq = infer_freq_from_formula(formulas)
            # distinguish clearly between formula-fed files and true empty metadata-only sheets
            status = "NEEDS_REFRESH"
            notes = f"mode={mode} formulas={len(formulas)} rows={ws_f.max_row} cols={ws_f.max_column}"

        candidate = {
            "sheet": ws_f.title,
            "header_row": header_row,
            "data_start_row": data_start_row,
            "freq": freq,
            "variables": variables,
            "status": status,
            "notes": notes,
            "engine": "openpyxl",
        }

        if best is None or candidate["data_start_row"] < best["data_start_row"]:
            best = candidate

    if best is None:
        # Try to identify sheets with metadata but no workable data.
        return {
            "sheet": wb_formula.sheetnames[0] if wb_formula.sheetnames else None,
            "header_row": None,
            "data_start_row": None,
            "freq": None,
            "variables": [],
            "status": "EMPTY_METADATA_ONLY",
            "notes": "no detected observation rows",
            "engine": "openpyxl",
        }

    return best


entries = []
for fp in FILES:
    rel = fp.relative_to(DATA_ROOT).as_posix()
    parts = rel.split("/")
    category = parts[0] if len(parts) > 0 else ""
    subcategory = parts[1] if len(parts) > 1 else ""
    ticker = fp.stem
    result = detect_sheet(fp)
    entries.append({
        "id": f"{category}_{subcategory}_{ticker}".replace(" ", "_").replace("-", "_").lower(),
        "path": f"data/{rel}",
        "category": category,
        "subcategory": subcategory,
        "ticker": ticker,
        "sheet": result["sheet"],
        "header_row": result["header_row"],
        "data_start_row": result["data_start_row"],
        "freq": result["freq"],
        "variables": result["variables"],
        "unit": infer_unit(rel),
        "status": result["status"],
        "notes": result["notes"],
    })

entries = sorted(entries, key=lambda x: x["path"])

payload = {
    "version": 1,
    "generated_at": pd.Timestamp.utcnow().isoformat(),
    "rules": {
        "header_row": "row immediately above the first observation row",
        "data_start_row": "first observation row with a cached date, or formula anchor row when Refinitiv values require refresh",
        "freq": "inferred from actual dates when available, otherwise from the RDP interval token",
        "variables": "raw column labels from the detected header row",
    },
    "datasets": entries,
}

with open(OUT_PATH, "w", encoding="utf-8") as f:
    yaml.safe_dump(payload, f, sort_keys=False, allow_unicode=True, width=140)

print(f"wrote {OUT_PATH}")
print(f"datasets={len(entries)}")
print(entries[0])
print(entries[-1])