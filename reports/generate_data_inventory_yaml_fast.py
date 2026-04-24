from pathlib import Path
import pandas as pd
import yaml
from openpyxl import load_workbook

ROOT = Path(r"c:/Users/Jorge/OneDrive - UFV/BAINF/PAPER")
DATA_ROOT = ROOT / "data"
OUT_PATH = ROOT / "reports" / "data_inventory.yml"
FILES = sorted(DATA_ROOT.rglob("*.xlsx"))

MANUAL = {
    "data/regime_variables/inflation/Germany_HICP.xlsx": {
        "sheet": "Historical Values",
        "header_row": 20,
        "data_start_row": 21,
        "freq": "monthly",
        "variables": ["Period", "HICP Prelim MM * (First Release)", "Poll", "Min", "Max", "HICP Final MM* (First Release)", "Poll", "Min", "Max"],
        "status": "NEEDS_REFRESH",
        "notes": "Refinitiv formula block; refresh in Excel to populate cached values",
    },
    "data/regime_variables/sentiment/ZEW_Germany.xlsx": {
        "sheet": "Historical Values",
        "header_row": 19,
        "data_start_row": 20,
        "freq": "monthly",
        "variables": ["Period", "aDEZEWSAR", "pDEZEWS=M", "pDEZEWS=L", "pDEZEWS=H"],
        "status": "NEEDS_REFRESH",
        "notes": "Refinitiv formula block; refresh in Excel to populate cached values",
    },
    "data/regime_variables/inflation/Eurozone_PPI.xlsx": {
        "sheet": "Historical Values",
        "header_row": None,
        "data_start_row": None,
        "freq": "monthly",
        "variables": [],
        "status": "EMPTY_METADATA_ONLY",
        "notes": "metadata only in current workbook",
    },
    "data/regime_variables/monetary/Germany_2Y_Yield.xlsx": {
        "sheet": "Historical Values",
        "header_row": None,
        "data_start_row": None,
        "freq": "monthly",
        "variables": [],
        "status": "EMPTY_METADATA_ONLY",
        "notes": "metadata only in current workbook",
    },
    "data/regime_variables/monetary/ECB_Balance_Sheet.xlsx": {
        "sheet": "Balance Sheet",
        "header_row": None,
        "data_start_row": None,
        "freq": None,
        "variables": [],
        "status": "WRONG_ENTITY",
        "notes": "company fundamentals workbook, not a timeseries",
    },
}


def norm(v):
    if v is None:
        return ""
    s = str(v).strip()
    return "" if s.lower() in {"none", "nan", "nat"} else s


def to_date(v):
    if v is None:
        return None
    if hasattr(v, "year") and hasattr(v, "month") and hasattr(v, "day"):
        return pd.Timestamp(v)
    s = norm(v)
    if not s:
        return None
    x = pd.to_datetime(s, errors="coerce", dayfirst=True)
    if pd.notna(x):
        return x
    x = pd.to_datetime(s, errors="coerce", dayfirst=False)
    if pd.notna(x):
        return x
    return None


def to_num(v):
    if isinstance(v, (int, float)):
        return float(v)
    s = norm(v).replace("%", "").replace("\u00a0", "")
    if not s:
        return None
    if s.count(",") == 1 and s.count(".") == 0:
        s = s.replace(",", ".")
    elif s.count(",") and s.count("."):
        if s.rfind(",") > s.rfind("."):
            s = s.replace(".", "").replace(",", ".")
        else:
            s = s.replace(",", "")
    try:
        return float(s)
    except Exception:
        return None


def infer_unit(rel_path):
    low = rel_path.lower()
    name = Path(rel_path).name.lower()
    if "/investable_assets/equity/" in low or "/investable_assets/commodities/" in low or "/investable_assets/real_assets/" in low:
        return "price_index_level"
    if "/investable_assets/cash/" in low:
        return "rate_level"
    if "/regime_variables/volatility/" in low:
        return "volatility_index_level"
    if "/regime_variables/sovereign_yields/" in low or "yield" in name:
        return "yield_level"
    if "hicp" in name or "ppi" in name:
        return "inflation_level"
    if "pmi" in name:
        return "diffusion_index"
    if "confidence" in name or "sentiment" in name or "zew" in name:
        return "sentiment_index"
    if "gdp" in name or "unemployment" in name or "industrial_production" in name:
        return "macro_level"
    if "balance_sheet" in name:
        return "balance_sheet_level"
    if "dxy" in name:
        return "fx_index_level"
    return "level"


def infer_freq(dates):
    series = pd.Series([pd.Timestamp(d) for d in dates if d is not None]).dropna().drop_duplicates().sort_values()
    if len(series) < 3:
        return "unknown"
    gaps = series.diff().dropna().dt.days
    if gaps.empty:
        return "unknown"
    med = float(gaps.median())
    if med <= 2:
        return "daily"
    if med <= 8:
        return "weekly"
    if med <= 16:
        return "biweekly"
    if med <= 40:
        return "monthly"
    if med <= 100:
        return "quarterly"
    if med <= 370:
        return "annual"
    return "irregular"


def clean_vars(values):
    out = []
    for value in values:
        s = norm(value)
        if s and not s.startswith("Unnamed:"):
            out.append(s)
    return out


entries = []
for fp in FILES:
    rel = fp.relative_to(ROOT).as_posix()
    key = f"data/{rel}"
    if key in MANUAL:
        m = MANUAL[key]
        entry = {
            "id": key.replace("data/", "").replace("/", "_").replace(" ", "_").replace("-", "_").replace(".xlsx", "").lower(),
            "path": key,
            "category": rel.split("/")[0],
            "subcategory": rel.split("/")[1] if len(rel.split("/")) > 1 else "",
            "ticker": fp.stem,
            "sheet": m["sheet"],
            "header_row": m["header_row"],
            "data_start_row": m["data_start_row"],
            "freq": m["freq"],
            "variables": m["variables"],
            "unit": infer_unit(rel),
            "status": m["status"],
            "notes": m["notes"],
        }
        entries.append(entry)
        continue

    try:
        wb = load_workbook(fp, data_only=True, read_only=True)
    except Exception as e:
        entries.append({
            "id": key.replace("data/", "").replace("/", "_").replace(" ", "_").replace("-", "_").replace(".xlsx", "").lower(),
            "path": key,
            "category": rel.split("/")[0],
            "subcategory": rel.split("/")[1] if len(rel.split("/")) > 1 else "",
            "ticker": fp.stem,
            "sheet": None,
            "header_row": None,
            "data_start_row": None,
            "freq": None,
            "variables": [],
            "unit": infer_unit(rel),
            "status": "UNREADABLE",
            "notes": f"cannot open workbook: {e}",
        })
        continue

    chosen = None
    for ws in wb.worksheets:
        for r in range(1, min(ws.max_row, 160) + 1):
            vals = [ws.cell(r, c).value for c in range(1, min(ws.max_column, 12) + 1)]
            if any(to_date(v) is not None for v in vals[:3]) and sum(to_num(v) is not None for v in vals[1:]) >= 1:
                header_row = r - 1
                while header_row > 1:
                    header_vals = [ws.cell(header_row, c).value for c in range(1, min(ws.max_column, 30) + 1)]
                    if any(norm(v) for v in header_vals):
                        break
                    header_row -= 1
                vars_ = clean_vars(header_vals)
                dates = []
                for rr in range(r, min(ws.max_row, r + 1000) + 1):
                    d = to_date(ws.cell(rr, 1).value)
                    if d is not None:
                        dates.append(d)
                chosen = {
                    "sheet": ws.title,
                    "header_row": header_row,
                    "data_start_row": r,
                    "freq": infer_freq(dates),
                    "variables": vars_,
                    "status": "OK",
                    "notes": f"mode=cached rows={ws.max_row} cols={ws.max_column}",
                }
                break
        if chosen:
            break

    if chosen is None:
        # Conservative fallback for other formula-driven or hidden sheets.
        entries.append({
            "id": key.replace("data/", "").replace("/", "_").replace(" ", "_").replace("-", "_").replace(".xlsx", "").lower(),
            "path": key,
            "category": rel.split("/")[0],
            "subcategory": rel.split("/")[1] if len(rel.split("/")) > 1 else "",
            "ticker": fp.stem,
            "sheet": wb.sheetnames[0] if wb.sheetnames else None,
            "header_row": None,
            "data_start_row": None,
            "freq": None,
            "variables": [],
            "unit": infer_unit(rel),
            "status": "EMPTY_METADATA_ONLY",
            "notes": "no cached observation rows found",
        })
        continue

    entries.append({
        "id": key.replace("data/", "").replace("/", "_").replace(" ", "_").replace("-", "_").replace(".xlsx", "").lower(),
        "path": key,
        "category": rel.split("/")[0],
        "subcategory": rel.split("/")[1] if len(rel.split("/")) > 1 else "",
        "ticker": fp.stem,
        "sheet": chosen["sheet"],
        "header_row": chosen["header_row"],
        "data_start_row": chosen["data_start_row"],
        "freq": chosen["freq"],
        "variables": chosen["variables"],
        "unit": infer_unit(rel),
        "status": chosen["status"],
        "notes": chosen["notes"],
    })

entries = sorted(entries, key=lambda x: x["path"])

payload = {
    "version": 1,
    "generated_at": pd.Timestamp.utcnow().isoformat(),
    "rules": {
        "header_row": "row immediately above the first observation row",
        "data_start_row": "first observation row with a cached date; manual overrides are used for formula-driven sheets that require Excel refresh",
        "freq": "inferred from cached dates when available",
        "variables": "raw column labels from the detected header row",
    },
    "datasets": entries,
}

with open(OUT_PATH, "w", encoding="utf-8") as f:
    yaml.safe_dump(payload, f, sort_keys=False, allow_unicode=True, width=140)

print(f"wrote {OUT_PATH}")
print(f"datasets={len(entries)}")
print(entries[0]["path"], entries[0]["status"], entries[0]["data_start_row"])
print(entries[-1]["path"], entries[-1]["status"], entries[-1]["data_start_row"])
